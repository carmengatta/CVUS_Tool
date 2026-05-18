"""Build a deferred-lives summary workbook from the extraction outputs.

Reads:  data_output/deferred_lives_top100_{year}.csv
        data_output/deferred_lives_payment_streams_{year}.csv
Writes: data_output/deferred_lives_summary_{year}.xlsx

Sheets:
  Overview               -- headline aggregates, industry rollup, data-quality summary
  Plan Detail            -- one row per plan, sorted by deferred-vested funding target
  TV Payment Streams     -- terminated-vested projected benefit payments, plans x years
  Methodology & Notes    -- column glossary, caveats, plans with no payment-stream data

Usage:
    python build_deferred_lives_excel.py --year 2024
"""

import argparse
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent

# ----- styles -----
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12, color="2F5496")
ITALIC = Font(italic=True, size=10)
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
NUM = "#,##0"
DOLLAR = "$#,##0"
DOLLAR_M = "$#,##0,,\" M\""
PCT = "0.0%"
FILL_HIGH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_MED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_LOW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_NA = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def cell(ws, row, col, value=None, align="left", fmt=None, border=True, font=None, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if border:
        c.border = THIN
    if fmt:
        c.number_format = fmt
    if font:
        c.font = font
    if fill:
        c.fill = fill
    return c


def auto_width(ws, min_w=10, max_w=48, only=None):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        if only is not None and col_cells[0].column not in only:
            continue
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[letter].width = min(max(longest + 2, min_w), max_w)


def _conf_fill(conf):
    return {"high": FILL_HIGH, "medium": FILL_MED, "low": FILL_LOW}.get(conf, FILL_NA)


# =============================================================================
# SHEET BUILDERS
# =============================================================================

def build_overview(wb, df, year):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = "2F5496"

    cell(ws, 1, 1, f"Deferred-Lives Summary — Top-100 DB Plans, {year} Filings", border=False, font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell(ws, 2, 1, "Terminated-vested ('deferred vested') participant data pulled from saved Form 5500 / "
                   "Schedule SB filings. See the 'Methodology & Notes' sheet.", border=False, font=ITALIC)

    n = len(df)
    have_stream = df["ebp_table_found"].sum()
    tot_ct = int(df["deferred_vested_count_sb"].sum())
    tot_ft = float(df["deferred_vested_funding_target_sb"].sum())
    tot_plan_ft = float(df["total_total_ft"].sum())

    r = 4
    cell(ws, r, 1, "Headline figures", border=False, font=SECTION_FONT); r += 1
    rows = [
        ("Plans covered", n, NUM),
        ("Deferred-vested participants (Schedule SB line 3b)", tot_ct, NUM),
        ("Deferred-vested funding target (Schedule SB line 3b)", tot_ft, DOLLAR),
        ("Deferred-vested share of total funding target (aggregate)", tot_ft / tot_plan_ft if tot_plan_ft else None, PCT),
        ("Median plan: deferred share of funding target", df["deferred_share_of_funding_target"].median(), PCT),
        ("Median plan: deferred share of headcount", df["deferred_share_of_headcount"].median(), PCT),
        ("Median plan: funding target per deferred life", df["avg_funding_target_per_deferred_life"].median(), DOLLAR),
        ("Plans with a line 26b projected-payment stream", f"{have_stream} of {n}", None),
        ("Total projected payments to deferred lives, next 10 yrs (where available)",
         float(df["tv_payments_next10"].dropna().sum()), DOLLAR),
    ]
    for label, val, fmt in rows:
        cell(ws, r, 1, label)
        cell(ws, r, 2, val, align="right", fmt=fmt)
        r += 1

    # ---- industry rollup ----
    r += 1
    cell(ws, r, 1, "By industry sector", border=False, font=SECTION_FONT); r += 1
    hdr = ["Industry sector", "Plans", "Deferred-vested participants",
           "Deferred-vested funding target", "Median deferred share of FT", "Median FT per deferred life"]
    for i, h in enumerate(hdr, 1):
        cell(ws, r, i, h)
    style_header(ws, r, len(hdr)); r += 1
    g = (df.groupby("INDUSTRY_SECTOR")
           .agg(plans=("EIN", "size"),
                ct=("deferred_vested_count_sb", "sum"),
                ft=("deferred_vested_funding_target_sb", "sum"),
                med_share=("deferred_share_of_funding_target", "median"),
                med_perlife=("avg_funding_target_per_deferred_life", "median"))
           .sort_values("ft", ascending=False))
    for sector, row in g.iterrows():
        cell(ws, r, 1, sector)
        cell(ws, r, 2, int(row["plans"]), align="center")
        cell(ws, r, 3, int(row["ct"]), align="right", fmt=NUM)
        cell(ws, r, 4, float(row["ft"]), align="right", fmt=DOLLAR)
        cell(ws, r, 5, row["med_share"], align="right", fmt=PCT)
        cell(ws, r, 6, row["med_perlife"], align="right", fmt=DOLLAR)
        r += 1
    cell(ws, r, 1, "Total", font=Font(bold=True))
    cell(ws, r, 2, int(g["plans"].sum()), align="center", font=Font(bold=True))
    cell(ws, r, 3, int(g["ct"].sum()), align="right", fmt=NUM, font=Font(bold=True))
    cell(ws, r, 4, float(g["ft"].sum()), align="right", fmt=DOLLAR, font=Font(bold=True))
    r += 2

    # ---- data quality ----
    cell(ws, r, 1, "Data quality — line 26b column identification", border=False, font=SECTION_FONT); r += 1
    for i, h in enumerate(["Outcome", "Plans", "Notes"], 1):
        cell(ws, r, i, h)
    style_header(ws, r, 3); r += 1
    method = df["tv_payment_col_method"].fillna("(no table)")
    conf = df["tv_payment_confidence"].fillna("(no table)")
    dq = [
        ("Stream parsed, column read from a clean header row", int((method == "header").sum()),
         "Highest confidence."),
        ("Stream parsed, column inferred (monotonic annuitant column + line 3 funding targets)",
         int((method == "heuristic").sum()),
         "Confidence flagged high / medium per plan in 'Plan Detail'."),
        ("No line 26b table — scanned-image attachment, no extractable text",
         int((df["ebp_table_found"] == False).sum()),  # noqa: E712
         "Would need OCR; line 3b headcount / funding target still available."),
    ]
    for label, cnt, note in dq:
        cell(ws, r, 1, label); cell(ws, r, 2, cnt, align="center"); cell(ws, r, 3, note); r += 1
    r += 1
    for label, cnt in [("Column-ID confidence: high", int((conf == "high").sum())),
                       ("Column-ID confidence: medium", int((conf == "medium").sum())),
                       ("Column-ID confidence: low", int((conf == "low").sum()))]:
        cell(ws, r, 1, label); cell(ws, r, 2, cnt, align="center"); r += 1

    auto_width(ws, max_w=60)
    ws.column_dimensions["A"].width = 62


PLAN_COLUMNS = [
    # (header, source column, align, number format)
    ("Sponsor", "SPONSOR_DFE_NAME", "left", None),
    ("Plan name", "PLAN_NAME", "left", None),
    ("Industry sector", "INDUSTRY_SECTOR", "left", None),
    ("EIN", "EIN", "center", None),
    ("PN", "PLAN_NUMBER", "center", None),
    ("Deferred-vested count (SB line 3b)", "deferred_vested_count_sb", "right", NUM),
    ("Deferred-vested count (5500 line 6)", "deferred_vested_count_5500", "right", NUM),
    ("Deferred-vested funding target", "deferred_vested_funding_target_sb", "right", DOLLAR),
    ("Deferred share of funding target", "deferred_share_of_funding_target", "right", PCT),
    ("Deferred share of headcount", "deferred_share_of_headcount", "right", PCT),
    ("Funding target per deferred life", "avg_funding_target_per_deferred_life", "right", DOLLAR),
    ("Retired count (SB line 3a)", "retired_count", "right", NUM),
    ("Active count (SB line 3c)", "active_count", "right", NUM),
    ("Total funding target (SB line 3d)", "total_total_ft", "right", DOLLAR),
    ("Annuitant + deferred share of FT", "annuitant_plus_deferred_share_ft", "right", PCT),
    ("TV projected payments, next 10 yrs", "tv_payments_next10", "right", DOLLAR),
    ("TV projected payments, full horizon", "tv_payments_total", "right", DOLLAR),
    ("Projection years", "ebp_n_years", "center", None),
    ("Assumed TV retirement age", "tv_retirement_age", "center", None),
    ("Actuary firm", "ACTUARY_FIRM_NAME", "left", None),
    ("Column-ID method", "tv_payment_col_method", "center", None),
    ("Column-ID confidence", "tv_payment_confidence", "center", None),
    ("ACK_ID (Form 5500)", "ACK_ID_5500", "left", None),
]


def build_plan_detail(wb, df):
    ws = wb.create_sheet("Plan Detail")
    ws.sheet_properties.tabColor = "70AD47"
    d = df.sort_values("deferred_vested_funding_target_sb", ascending=False, na_position="last").reset_index(drop=True)

    cell(ws, 1, 1, "Deferred-Lives Detail by Plan (sorted by deferred-vested funding target)", border=False, font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    hdr_row = 3
    for i, (h, *_rest) in enumerate(PLAN_COLUMNS, 1):
        cell(ws, hdr_row, i, h)
    style_header(ws, hdr_row, len(PLAN_COLUMNS))

    conf_col = next(i for i, (h, *_r) in enumerate(PLAN_COLUMNS, 1) if h == "Column-ID confidence")
    for ridx, row in d.iterrows():
        r = hdr_row + 1 + ridx
        for cidx, (h, src, align, fmt) in enumerate(PLAN_COLUMNS, 1):
            val = row.get(src)
            if pd.isna(val):
                val = None
            elif fmt == NUM or h in ("Projection years", "Assumed TV retirement age"):
                val = int(val) if isinstance(val, float) and float(val).is_integer() else val
            cell(ws, r, cidx, val, align=align, fmt=fmt)
        # colour the confidence cell
        ws.cell(row=r, column=conf_col).fill = _conf_fill(row.get("tv_payment_confidence"))

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=3)
    auto_width(ws, max_w=42)
    # keep the first two text columns readable but bounded
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["T"].width = 26  # actuary firm


def build_payment_streams(wb, df, streams, year, horizon=30):
    ws = wb.create_sheet("TV Payment Streams")
    ws.sheet_properties.tabColor = "FFC000"

    cell(ws, 1, 1, f"Projected Benefit Payments to Terminated-Vested Participants — Schedule SB line 26b "
                   f"({year}–{year + horizon - 1})", border=False, font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    cell(ws, 2, 1, "One row per plan. Blank = no line 26b table in the saved filing (scanned attachment). "
                   "Confidence reflects how the terminated-vested column was identified — see 'Methodology & Notes'.",
         border=False, font=ITALIC)

    years = list(range(year, year + horizon))
    pv = (streams[streams["year"].isin(years)]
          .pivot_table(index=["EIN", "PLAN_NUMBER"], columns="year", values="term_vested", aggfunc="first"))

    base = df.sort_values("deferred_vested_funding_target_sb", ascending=False, na_position="last")
    lead = ["Sponsor", "Plan name", "EIN", "PN", "Deferred-vested funding target", "Confidence"]
    hdr_row = 4
    for i, h in enumerate(lead, 1):
        cell(ws, hdr_row, i, h)
    for j, y in enumerate(years):
        cell(ws, hdr_row, len(lead) + 1 + j, y)
    style_header(ws, hdr_row, len(lead) + len(years))

    for ridx, row in base.reset_index(drop=True).iterrows():
        r = hdr_row + 1 + ridx
        cell(ws, r, 1, row["SPONSOR_DFE_NAME"])
        cell(ws, r, 2, row["PLAN_NAME"])
        cell(ws, r, 3, row["EIN"], align="center")
        cell(ws, r, 4, row["PLAN_NUMBER"], align="center")
        cell(ws, r, 5, None if pd.isna(row["deferred_vested_funding_target_sb"]) else float(row["deferred_vested_funding_target_sb"]),
             align="right", fmt=DOLLAR)
        cell(ws, r, 6, row.get("tv_payment_confidence"), align="center", fill=_conf_fill(row.get("tv_payment_confidence")))
        key = (str(row["EIN"]).zfill(9), int(row["PLAN_NUMBER"]))
        for j, y in enumerate(years):
            c = cell(ws, r, len(lead) + 1 + j, None, align="right", fmt=DOLLAR)
            try:
                v = pv.loc[key, y]
                if pd.notna(v):
                    c.value = float(v)
            except (KeyError, TypeError):
                pass

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=len(lead) + 1)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12
    for j in range(len(years)):
        ws.column_dimensions[get_column_letter(len(lead) + 1 + j)].width = 15


TV_ASSUMPTION_COLUMNS = [
    ("Sponsor", "SPONSOR_DFE_NAME", "left", None),
    ("Plan name", "PLAN_NAME", "left", None),
    ("Industry sector", "INDUSTRY_SECTOR", "left", None),
    ("Actuary firm", "ACTUARY_FIRM_NAME", "left", None),
    ("EIN", "EIN", "center", None),
    ("PN", "PLAN_NUMBER", "center", None),
    ("Deferred-vested funding target", "TERM_LIABILITY", "right", DOLLAR),
    ("SB attachments readable?", "has_extractable_sb_text", "center", None),
    ("# TV evidence blocks", "n_tv_evidence_blocks", "center", None),
    ("Assumption categories present", "categories_present", "left", None),
    ("Assumed TV commencement age", "tv_commencement_age", "center", None),
    ("…immediate if past that age", "tv_immediate_if_past_age", "center", None),
    ("Lump sum available to TV?", "tv_lump_sum_available", "center", None),
    ("TV lump-sum / annuity election", "tv_lump_sum_election", "left", None),
    ("TV normal form of payment", "tv_normal_form", "left", None),
    ("% married", "married_pct", "center", None),
    ("Spouse age-difference note", "spouse_age_diff_note", "left", None),
    ("Lump-sum / 417(e) basis flag", "lump_sum_417e_basis", "center", None),
    ("Cash-balance / PPA-VPA component?", "has_cash_balance_or_ppa_vpa", "center", None),
    ("Cash-balance crediting rate (note)", "cash_balance_crediting_rate", "left", None),
    ("Mortality table (rough)", "mortality_table", "center", None),
    ("Generational mortality?", "mortality_generational", "center", None),
    ("TV commencement-age note", "tv_commencement_age_note", "left", None),
    ("ACK_ID (Form 5500)", "ACK_ID_5500", "left", None),
]


def _yn_fill(val):
    if val is True or (isinstance(val, str) and val.strip().lower() in ("true", "yes")):
        return FILL_HIGH
    if val is False or (isinstance(val, str) and val.strip().lower() in ("false", "no")):
        return FILL_NA
    return None


def build_tv_assumptions(wb, tv):
    ws = wb.create_sheet("TV Assumptions")
    ws.sheet_properties.tabColor = "7030A0"
    d = tv.sort_values(["has_extractable_sb_text", "TERM_LIABILITY"], ascending=[False, False], na_position="last").reset_index(drop=True)

    cell(ws, 1, 1, "Terminated-Vested Actuarial Assumptions — Structured Summary by Plan", border=False, font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell(ws, 2, 1, "Best-effort parse of the SB-attachment assumptions prose — many cells are blank where the filing "
                   "doesn't state it cleanly. The full source text is on the 'TV Assumptions — Evidence' sheet "
                   "(one row per passage). See 'Methodology & Notes'.", border=False, font=ITALIC)

    hdr_row = 4
    for i, (h, *_r) in enumerate(TV_ASSUMPTION_COLUMNS, 1):
        cell(ws, hdr_row, i, h)
    style_header(ws, hdr_row, len(TV_ASSUMPTION_COLUMNS))

    readable_col = next(i for i, (h, *_r) in enumerate(TV_ASSUMPTION_COLUMNS, 1) if h == "SB attachments readable?")
    yn_cols = {i for i, (h, src, *_r) in enumerate(TV_ASSUMPTION_COLUMNS, 1)
               if src in ("tv_lump_sum_available", "lump_sum_417e_basis", "has_cash_balance_or_ppa_vpa",
                          "tv_immediate_if_past_age", "mortality_generational", "has_extractable_sb_text")}

    for ridx, row in d.iterrows():
        r = hdr_row + 1 + ridx
        for cidx, (h, src, align, fmt) in enumerate(TV_ASSUMPTION_COLUMNS, 1):
            val = row.get(src)
            if pd.isna(val):
                val = None
            elif h in ("Assumed TV commencement age", "% married", "# TV evidence blocks"):
                try:
                    val = int(val)
                except (TypeError, ValueError):
                    pass
            c = cell(ws, r, cidx, val, align=align, fmt=fmt)
            if cidx in yn_cols:
                f = _yn_fill(row.get(src))
                if f:
                    c.fill = f

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=3)
    auto_width(ws, max_w=40)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 26
    # wrap the long free-text columns
    for col_idx, (h, src, *_r) in enumerate(TV_ASSUMPTION_COLUMNS, 1):
        if src in ("categories_present", "tv_lump_sum_election", "tv_normal_form", "spouse_age_diff_note",
                   "cash_balance_crediting_rate", "tv_commencement_age_note"):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = 46
            for r in range(hdr_row + 1, hdr_row + 1 + len(d)):
                ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")


def build_tv_evidence(wb, evidence):
    ws = wb.create_sheet("TV Assumptions — Evidence")
    ws.sheet_properties.tabColor = "C586E5"
    cell(ws, 1, 1, "Terminated-Vested Assumptions — Source Passages", border=False, font=TITLE_FONT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    cell(ws, 2, 1, "Every passage in the saved SB attachments that mentions a terminated-vested / deferred-vested "
                   "participant (line 3b and line 26b boilerplate excluded), with the assumption categories it touches. "
                   "Filter by Sponsor or Categories to pull a plan's full set of TV-related assumption text.",
         border=False, font=ITALIC)

    cols = [("Sponsor", "SPONSOR_DFE_NAME", 34), ("Plan name", "PLAN_NAME", 38), ("EIN", "EIN", 12),
            ("PN", "PLAN_NUMBER", 6), ("Block #", "block_no", 8), ("Categories", "categories", 30),
            ("Source text", "text", 140)]
    hdr_row = 4
    for i, (h, *_r) in enumerate(cols, 1):
        cell(ws, hdr_row, i, h)
    style_header(ws, hdr_row, len(cols))

    ev = evidence.sort_values(["SPONSOR_DFE_NAME", "PLAN_NUMBER", "block_no"]).reset_index(drop=True)
    for ridx, row in ev.iterrows():
        r = hdr_row + 1 + ridx
        for cidx, (h, src, _w) in enumerate(cols, 1):
            val = row.get(src)
            if pd.isna(val):
                val = None
            elif h == "Block #":
                try:
                    val = int(val)
                except (TypeError, ValueError):
                    pass
            c = cell(ws, r, cidx, val, align="left")
            if src == "text":
                c.alignment = Alignment(wrap_text=True, vertical="top")
    for i, (_h, _src, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(cols))}{hdr_row + len(ev)}"


def build_notes(wb, df, tv=None):
    ws = wb.create_sheet("Methodology & Notes")
    ws.sheet_properties.tabColor = "808080"
    cell(ws, 1, 1, "Methodology & Notes", border=False, font=TITLE_FONT)

    notes = [
        ("Source", "Pulled from the saved top-100 Form 5500 / Schedule SB filings "
                   "(pdf_extraction/form5500_chunked_2024/). Plan identifiers, industry, total funding target and the "
                   "Form 5500 line 6 participant counts are joined from data_output/yearly/db_plans_2024.parquet."),
        ("'Deferred lives'", "Terminated-vested participants — people who left the employer with a vested benefit but "
                             "have not yet started receiving it. On Schedule SB they are line 3b; on Form 5500 line 6 "
                             "they are the 'separated participants entitled to future benefits' count."),
        ("Deferred-vested count (SB line 3b)", "Headcount from Schedule SB, Part I, line 3b. Parsed and self-validated "
                                               "(the line 3 a/b/c rows add to the line 3d total) for every saved filing."),
        ("Deferred-vested count (5500 line 6)", "Form 5500 line 6 'separated, future benefits' count. Typically runs a "
                                                "bit below the SB line 3b figure — different measurement date and basis. "
                                                "Both are shown; neither is 'wrong'."),
        ("Deferred-vested funding target", "Schedule SB line 3b total funding target (= the vested funding target — "
                                           "terminated-vested participants are by definition fully vested). Matches the "
                                           "TERM_LIABILITY field in the master parquet."),
        ("TV projected payments (next 10 yrs / full horizon)", "Sum of the 'Terminated Vested' column of the Schedule SB "
                                                               "line 26b 'Schedule of Projection of Expected Benefit "
                                                               "Payments' attachment. Undiscounted projected cash flows."),
        ("Column-ID method / confidence", "The line 26b table has Active / Terminated-Vested / Retired columns whose "
                                          "labels are often jumbled by PDF text extraction. 'header' = the column-name "
                                          "row was rendered cleanly and read directly (high confidence). 'heuristic' = "
                                          "the annuitant column was identified by its monotonic decline and the active / "
                                          "terminated-vested columns were told apart using the line 3 funding targets; "
                                          "confidence is high when those funding targets differ clearly, medium otherwise. "
                                          "'low' / blank = could not be verified or no table present."),
        ("Assumed TV retirement age", "Best-effort extraction from the SB-attachment assumptions prose where stated; "
                                      "blank for plans that don't state it in a machine-readable way."),
        ("Plans with no line 26b table", "These ship their SB attachments as scanned images with no extractable text, so "
                                         "the payment stream could not be parsed (line 3b headcount and funding target "
                                         "are still available). Listed below — they would need OCR."),
        ("'TV Assumptions' sheet", "Structured, best-effort parse of the terminated-vested-related actuarial assumptions "
                                   "in the SB-attachment prose: assumed benefit-commencement / retirement age, whether a "
                                   "lump sum is available to deferred vesteds and the election split, normal form of "
                                   "payment, married %, spouse age difference, whether a lump-sum / 417(e) basis is "
                                   "referenced, whether the plan has a cash-balance / PPA-VPA component and its crediting "
                                   "rate, and a rough mortality read. The prose varies a lot by actuarial firm, so many "
                                   "cells are blank — treat this sheet as a quick filter, not the final word."),
        ("'TV Assumptions — Evidence' sheet", "The workhorse: every passage in the saved SB attachments that mentions a "
                                              "terminated-vested / deferred-vested participant (line 3b and line 26b "
                                              "boilerplate excluded), one row per passage, tagged with the assumption "
                                              "categories it touches (commencement age, form of payment, lump-sum basis, "
                                              "spouse, mortality, cash-balance/PPA-VPA, death benefit, expenses, data "
                                              "treatment). Filter by Sponsor to read a plan's complete TV assumption text. "
                                              "Only the 77 of 99 plans whose SB attachments have extractable text are "
                                              "covered; the other 22 are scanned and would need OCR."),
    ]
    r = 3
    for label, body in notes:
        cell(ws, r, 1, label, font=Font(bold=True))
        cell(ws, r, 2, body)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1
    cell(ws, r, 1, "Plans with no line 26b payment stream:", font=Font(bold=True)); r += 1
    for _, row in df[df["ebp_table_found"] == False].sort_values("deferred_vested_funding_target_sb", ascending=False).iterrows():  # noqa: E712
        cell(ws, r, 1, f"{row['SPONSOR_DFE_NAME']} (EIN {row['EIN']}, PN {row['PLAN_NUMBER']})")
        cell(ws, r, 2, f"deferred-vested funding target ${row['deferred_vested_funding_target_sb']:,.0f}")
        r += 1
    if tv is not None:
        r += 1
        cell(ws, r, 1, "Plans with no extractable SB-attachment text (TV assumptions need OCR):", font=Font(bold=True)); r += 1
        for _, row in tv[tv["has_extractable_sb_text"] == False].sort_values("TERM_LIABILITY", ascending=False, na_position="last").iterrows():  # noqa: E712
            cell(ws, r, 1, f"{row['SPONSOR_DFE_NAME']} (EIN {row['EIN']}, PN {row['PLAN_NUMBER']})")
            tl = row.get("TERM_LIABILITY")
            cell(ws, r, 2, f"deferred-vested funding target ${tl:,.0f}" if pd.notna(tl) else "")
            r += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 115
    for row_cells in ws.iter_rows():
        for c in row_cells:
            if c.column == 2:
                c.alignment = Alignment(wrap_text=True, vertical="top")


# =============================================================================
# MAIN
# =============================================================================

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--year", type=int, default=2024)
    ap.add_argument("--out-dir", type=Path, default=REPO_ROOT / "data_output")
    args = ap.parse_args()

    detail_csv = args.out_dir / f"deferred_lives_top100_{args.year}.csv"
    streams_csv = args.out_dir / f"deferred_lives_payment_streams_{args.year}.csv"
    if not detail_csv.exists():
        raise SystemExit(f"Missing {detail_csv} — run pdf_extraction/extract_deferred_lives.py first.")
    df = pd.read_csv(detail_csv)
    df["EIN"] = df["EIN"].astype(str).str.zfill(9)
    df["PLAN_NUMBER"] = df["PLAN_NUMBER"].astype(str).str.zfill(3)
    streams = pd.read_csv(streams_csv) if streams_csv.exists() else pd.DataFrame(
        columns=["EIN", "PLAN_NUMBER", "year", "term_vested"])
    if not streams.empty:
        streams["EIN"] = streams["EIN"].astype(str).str.zfill(9)

    # Terminated-vested assumptions (optional — produced by extract_tv_assumptions.py).
    tv_csv = args.out_dir / f"tv_assumptions_top100_{args.year}.csv"
    tv_ev_csv = args.out_dir / f"tv_assumptions_evidence_{args.year}.csv"
    tv = pd.read_csv(tv_csv) if tv_csv.exists() else None
    tv_ev = pd.read_csv(tv_ev_csv) if tv_ev_csv.exists() else None
    if tv is not None:
        tv["EIN"] = tv["EIN"].astype(str).str.zfill(9)
        tv["PLAN_NUMBER"] = tv["PLAN_NUMBER"].astype(str).str.zfill(3)
    if tv_ev is not None:
        tv_ev["EIN"] = tv_ev["EIN"].astype(str).str.zfill(9)
        tv_ev["PLAN_NUMBER"] = tv_ev["PLAN_NUMBER"].astype(str).str.zfill(3)

    wb = openpyxl.Workbook()
    build_overview(wb, df, args.year)
    build_plan_detail(wb, df)
    build_payment_streams(wb, df, streams, args.year)
    if tv is not None:
        build_tv_assumptions(wb, tv)
    if tv_ev is not None and not tv_ev.empty:
        build_tv_evidence(wb, tv_ev)
    build_notes(wb, df, tv)

    out_path = args.out_dir / f"deferred_lives_summary_{args.year}.xlsx"
    wb.save(out_path)
    print(f"Saved {out_path}")
    if tv is None:
        print("  (note: TV-assumptions sheets skipped — run pdf_extraction/extract_tv_assumptions.py to include them)")


if __name__ == "__main__":
    main()
