"""
Build a manual-review workbook for the parsed mortality table assignments.

For each plan:
  - shows the parsed fields (base_table, collar, scaling, improvement_scale, projection_method, weighting, company_adjusted, substitute_mortality)
  - shows the original mortality language from the Form 5500 .txt file
  - includes ~500 chars of context before/after each mortality mention (≈ 2-3 sentences)
  - restricts to Financial Statements pages only (Sch SB form-checkbox text is excluded)

Output: pdf_extraction/form5500_txt_2024/_mortality_review_{year}.xlsx

Usage:
    python pdf_extraction/build_mortality_review.py --year 2024
"""

import argparse
import re
from pathlib import Path

import pandas as pd

CONTEXT_BEFORE = 500
CONTEXT_AFTER = 500
SNIPPET_SEP = "\n\n--------------------------------\n\n"


def split_pages(text: str) -> dict:
    """Parse '--- PAGE N ---' boundaries into a {page_num: page_text} dict."""
    pages = {}
    parts = re.split(r'\n\s*--- PAGE (\d+) ---\s*\n', text)
    # parts = [pre, '1', text1, '2', text2, ...]
    for i in range(1, len(parts), 2):
        pages[int(parts[i])] = parts[i + 1]
    return pages


def clean_pdf_artifacts(text: str) -> str:
    """Strip pdfplumber/PDF noise that hurts readability without losing meaning."""
    text = re.sub(r'\(cid:\d+\)', '•', text)        # bullet glyphs encoded as cid
    text = re.sub(r'[ \t]+', ' ', text)              # collapse runs of spaces/tabs
    text = re.sub(r' *\n *', '\n', text)             # trim each line
    text = re.sub(r'\n{3,}', '\n\n', text)           # collapse blank lines
    return text.strip()


def extract_context_for_page(page_text: str, before: int = CONTEXT_BEFORE, after: int = CONTEXT_AFTER) -> list:
    """Find every 'mortality' match on a page and return a list of (start_idx, snippet_with_context)."""
    snippets = []
    seen_spans = []
    for m in re.finditer(r'mortality', page_text, re.IGNORECASE):
        s = max(0, m.start() - before)
        e = min(len(page_text), m.end() + after)
        # Dedupe overlapping windows: if this window overlaps a prior one by >50%, skip
        if any(not (e <= ps or s >= pe) and min(e, pe) - max(s, ps) > 0.5 * (e - s)
               for ps, pe in seen_spans):
            continue
        seen_spans.append((s, e))
        snippets.append((s, page_text[s:e].strip()))
    return snippets


def build_plan_review(year: int) -> pd.DataFrame:
    script_dir = Path(__file__).parent
    txt_dir = script_dir / f"form5500_txt_{year}"
    summary_path = txt_dir / f"_plan_mortality_summary_{year}.csv"
    metadata_path = txt_dir / f"_extraction_metadata_{year}.csv"

    summary = pd.read_csv(summary_path, dtype={'ein': str, 'plan_number': str})
    metadata = pd.read_csv(metadata_path)

    # Map pdf_file -> list of FS mortality pages
    fs_pages_map = {}
    for _, r in metadata.iterrows():
        pages_str = r.get('fs_mortality_pages')
        if pd.isna(pages_str) or not pages_str:
            continue
        fs_pages_map[r['pdf_file']] = [int(p.strip()) for p in str(pages_str).split(',') if p.strip()]

    # Try to attach sponsor name + participant count from the 2024 db parquet
    db_path = script_dir.parent / 'data_output' / 'yearly' / f'db_plans_{year}.parquet'
    sponsor_lookup = {}
    if db_path.exists():
        db = pd.read_parquet(db_path, columns=['EIN', 'PLAN_NUMBER', 'SPONSOR_DFE_NAME',
                                                'TOTAL_PARTICIPANTS', 'RETIREE_COUNT',
                                                'MORTALITY_CODE', 'ACTUARY_FIRM_NAME', 'INDUSTRY_SECTOR'])
        db['EIN'] = db['EIN'].astype(str).str.zfill(9)
        db['PLAN_NUMBER'] = db['PLAN_NUMBER'].astype(str).str.zfill(3)
        for _, r in db.iterrows():
            sponsor_lookup[(r['EIN'], r['PLAN_NUMBER'])] = (
                r['SPONSOR_DFE_NAME'], r['TOTAL_PARTICIPANTS'], r['RETIREE_COUNT'],
                r['MORTALITY_CODE'], r['ACTUARY_FIRM_NAME'], r['INDUSTRY_SECTOR'],
            )

    rows = []
    for _, plan in summary.iterrows():
        pdf_file = plan['pdf_file']
        ein = str(plan['ein']).zfill(9)
        pn = str(plan['plan_number']).zfill(3)
        txt_file = txt_dir / pdf_file.replace('.pdf', '.txt')

        if not txt_file.exists():
            language = "(text file not found)"
            n_snippets = 0
        else:
            text = txt_file.read_text(encoding='utf-8', errors='replace')
            pages = split_pages(text)
            fs_pages = fs_pages_map.get(pdf_file, [])

            all_snippets = []
            for pg in fs_pages:
                if pg not in pages:
                    continue
                page_text = pages[pg]
                snippets = extract_context_for_page(page_text)
                for _, snip in snippets:
                    all_snippets.append(f"[PAGE {pg}]\n{clean_pdf_artifacts(snip)}")

            n_snippets = len(all_snippets)
            language = SNIPPET_SEP.join(all_snippets) if all_snippets else "(no FS mortality language found)"

        # Excel cell limit is 32,767 chars
        if len(language) > 32000:
            language = language[:31900] + "\n\n... [TRUNCATED — open .txt file for full text]"

        sponsor_info = sponsor_lookup.get((ein, pn), (None, None, None, None, None, None))

        rows.append({
            'sponsor': sponsor_info[0],
            'ein': ein,
            'plan_number': pn,
            'industry': sponsor_info[5],
            'total_participants': sponsor_info[1],
            'retirees': sponsor_info[2],
            'sb_mortality_code': sponsor_info[3],
            'actuary_firm': sponsor_info[4],
            'pdf_file': pdf_file,
            'PARSED_base_table': plan.get('base_table'),
            'PARSED_collar': plan.get('collar_adjustment'),
            'PARSED_scaling_factor': plan.get('scaling_factor'),
            'PARSED_scaling_description': plan.get('scaling_description'),
            'PARSED_improvement_scale': plan.get('improvement_scale'),
            'PARSED_projection_method': plan.get('projection_method'),
            'PARSED_projection_base_year': plan.get('projection_base_year'),
            'PARSED_weighting': plan.get('weighting'),
            'PARSED_company_adjusted': plan.get('company_adjusted'),
            'PARSED_substitute_mortality': plan.get('substitute_mortality'),
            'PARSED_covid_adjustment': plan.get('covid_adjustment'),
            'CORRECTED_base_table': '',
            'CORRECTED_collar': '',
            'CORRECTED_scaling_factor': '',
            'CORRECTED_improvement_scale': '',
            'CORRECTED_projection_method': '',
            'CORRECTED_weighting': '',
            'review_notes': '',
            'snippet_count': n_snippets,
            'mortality_language': language,
        })

    return pd.DataFrame(rows).sort_values(['industry', 'total_participants'],
                                            ascending=[True, False],
                                            na_position='last').reset_index(drop=True)


def build_stats(summary_path: Path, db_path: Path = None) -> pd.DataFrame:
    """Compute overall distributions used for the Stats sheet."""
    summary = pd.read_csv(summary_path, dtype={'ein': str, 'plan_number': str})
    if db_path and db_path.exists():
        db = pd.read_parquet(db_path, columns=['EIN', 'PLAN_NUMBER', 'TOTAL_PARTICIPANTS'])
        db['EIN'] = db['EIN'].astype(str).str.zfill(9)
        db['PLAN_NUMBER'] = db['PLAN_NUMBER'].astype(str).str.zfill(3)
        summary['ein'] = summary['ein'].astype(str).str.zfill(9)
        summary['plan_number'] = summary['plan_number'].astype(str).str.zfill(3)
        summary = summary.merge(db, left_on=['ein', 'plan_number'], right_on=['EIN', 'PLAN_NUMBER'], how='left')

    out = []

    def block(label, series, weights=None):
        out.append({'metric': label, 'value': '', 'plans': '', 'participants': ''})
        vc = series.fillna('(unparsed/none)').value_counts(dropna=False)
        for k, v in vc.items():
            participants = ''
            if weights is not None:
                participants = int(weights[series.fillna('(unparsed/none)') == k].sum() or 0)
            out.append({'metric': '', 'value': k, 'plans': int(v), 'participants': participants})
        out.append({'metric': '', 'value': '', 'plans': '', 'participants': ''})

    weights = summary.get('TOTAL_PARTICIPANTS')
    block('Base Table', summary['base_table'], weights)
    block('Improvement Scale', summary['improvement_scale'], weights)
    block('Projection Method', summary['projection_method'], weights)
    block('Collar Adjustment', summary['collar_adjustment'], weights)
    block('Weighting', summary['weighting'], weights)
    block('Company Adjusted (any sponsor mod)', summary['company_adjusted'], weights)
    block('Substitute Mortality (current)', summary['substitute_mortality'], weights)
    block('Former Substitute Mortality', summary['former_substitute_mortality'], weights)
    block('COVID Adjustment', summary['covid_adjustment'], weights)

    # Scaling factor stats
    sf = summary['scaling_factor'].dropna()
    out.append({'metric': 'Scaling Factor — count present', 'value': len(sf),
                'plans': len(summary), 'participants': ''})
    if len(sf):
        for stat in ['min', 'mean', 'median', 'max']:
            v = getattr(sf, stat)() if stat != 'median' else sf.median()
            out.append({'metric': f'Scaling Factor — {stat}', 'value': round(v, 4),
                        'plans': '', 'participants': ''})

    return pd.DataFrame(out)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--year', type=int, default=2024)
    args = p.parse_args()

    script_dir = Path(__file__).parent
    txt_dir = script_dir / f"form5500_txt_{args.year}"
    summary_path = txt_dir / f"_plan_mortality_summary_{args.year}.csv"
    db_path = script_dir.parent / 'data_output' / 'yearly' / f'db_plans_{args.year}.parquet'
    out_path = txt_dir / f"_mortality_review_{args.year}.xlsx"

    print(f"Building plan review from {summary_path}...")
    review = build_plan_review(args.year)
    print(f"  {len(review)} plans")

    print("Computing stats...")
    stats = build_stats(summary_path, db_path)

    # Build a "Gaps Only" sheet: plans missing one of the key parsed fields
    has_gap = (review['PARSED_base_table'].isna() |
               review['PARSED_collar'].isna() |
               review['PARSED_projection_method'].isna() |
               review['PARSED_improvement_scale'].isna())
    gap_review = review[has_gap].copy()

    def _gap_reason(r):
        missing = []
        if pd.isna(r['PARSED_base_table']):        missing.append('base_table')
        if pd.isna(r['PARSED_collar']):            missing.append('collar')
        if pd.isna(r['PARSED_projection_method']): missing.append('projection_method')
        if pd.isna(r['PARSED_improvement_scale']): missing.append('improvement_scale')
        return ', '.join(missing)

    gap_review['gap_reason'] = gap_review.apply(_gap_reason, axis=1)
    # Move gap_reason to the front
    cols = ['gap_reason'] + [c for c in gap_review.columns if c != 'gap_reason']
    gap_review = gap_review[cols]

    print(f"Writing {out_path}...")
    with pd.ExcelWriter(out_path, engine='openpyxl') as xl:
        stats.to_excel(xl, sheet_name='Stats', index=False)
        review.to_excel(xl, sheet_name='Plan Review', index=False)
        gap_review.to_excel(xl, sheet_name='Gaps Only', index=False)

    # Apply column widths + wrap on the language column
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter, column_index_from_string
    wb = load_workbook(out_path)

    # Stats sheet
    ws = wb['Stats']
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16

    # Apply formatting to both review sheets
    review_sheets = ['Plan Review', 'Gaps Only']
    widths = {
        'A': 36,  # sponsor
        'B': 12,  # ein
        'C': 8,   # pn
        'D': 22,  # industry
        'E': 14,  # participants
        'F': 12,  # retirees
        'G': 8,   # sb_code
        'H': 22,  # actuary
        'I': 38,  # pdf_file
        'J': 14, 'K': 14, 'L': 12, 'M': 24, 'N': 22, 'O': 16, 'P': 12, 'Q': 14, 'R': 14, 'S': 14, 'T': 12,  # PARSED_*
        'U': 14, 'V': 14, 'W': 12, 'X': 22, 'Y': 14, 'Z': 14,  # CORRECTED_*
        'AA': 28,  # review_notes
        'AB': 8,   # snippet_count
        'AC': 100, # mortality_language
    }
    wrap = Alignment(wrap_text=True, vertical='top')

    for sheet_name in review_sheets:
        ws = wb[sheet_name]
        df_for_sheet = gap_review if sheet_name == 'Gaps Only' else review
        # The Gaps sheet has gap_reason inserted as col A, so widths shift right by one
        col_offset = 1 if sheet_name == 'Gaps Only' else 0
        if col_offset:
            ws.column_dimensions['A'].width = 24  # gap_reason
        for col_letter, w in widths.items():
            shifted = get_column_letter(column_index_from_string(col_letter) + col_offset)
            ws.column_dimensions[shifted].width = w

        lang_col_idx = list(df_for_sheet.columns).index('mortality_language') + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=lang_col_idx).alignment = wrap
            ws.row_dimensions[row].height = 200
        # Freeze so the parsed columns stay visible while scrolling the language
        ws.freeze_panes = 'J2' if sheet_name == 'Plan Review' else 'K2'

    wb.save(out_path)
    print(f"Done: {out_path}")


if __name__ == '__main__':
    main()
