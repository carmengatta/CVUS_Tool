"""Build substitute_mortality_analysis.xlsx from the markdown analysis."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12, color="2F5496")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
num_fmt = '#,##0'
dollar_fmt = '$#,##0'


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border


def style_data_cell(ws, row, col, align="left"):
    cell = ws.cell(row=row, column=col)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal=align)
    return cell


def auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ===== SHEET 1: Summary =====
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_properties.tabColor = "2F5496"

ws1.cell(row=1, column=1, value="Substitute Mortality Table Analysis (2019-2024)").font = title_font
ws1.merge_cells("A1:D1")

# Year-by-Year Counts
ws1.cell(row=3, column=1, value="Plans Using Substitute Mortality by Year").font = section_font

headers = ["Plan Year", "Plans", "Participants", "YoY Change (Plans)"]
r = 4
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header_row(ws1, r, 4)

year_data = [
    (2019, 101, 1892958, None),
    (2020, 102, 1790895, 1),
    (2021, 104, 1765425, 2),
    (2022, 88, 1587297, -16),
    (2023, 83, 1510829, -5),
    (2024, 55, 950053, -28),
]
for i, (yr, plans, parts, chg) in enumerate(year_data):
    r = 5 + i
    style_data_cell(ws1, r, 1, "center").value = yr
    style_data_cell(ws1, r, 2, "center").value = plans
    c = style_data_cell(ws1, r, 3, "right")
    c.value = parts
    c.number_format = num_fmt
    style_data_cell(ws1, r, 4, "center").value = chg if chg is not None else "--"

r += 1
ws1.cell(row=r + 1, column=1,
         value="Net decline 2019-2024: 46 plans (45.5%), ~943,000 participants (49.8%)").font = Font(italic=True)

# Disposition Summary
r += 3
ws1.cell(row=r, column=1, value="Disposition of All 137 Plans (2019-2024)").font = section_font

r += 1
for i, h in enumerate(["Outcome", "Plans", "Participants"], 1):
    ws1.cell(row=r, column=i, value=h)
style_header_row(ws1, r, 3)

disp_data = [
    ("Still Substitute (active)", 55, 950053),
    ("Switched to Prescribed (still active)", 38, 195189),
    ("Terminated (final filing, no PRT)", 19, 53751),
    ("Terminated with PRT", 1, 921),
    ("PRT Activity (winding down)", 5, 15758),
    ("Plan Merger (EIN still active)", 6, 43833),
    ("Likely Late Filer (still Code 3)", 4, 399312),
    ("No Longer Filing (status unclear)", 9, 36542),
]

for i, (outcome, plans, parts) in enumerate(disp_data):
    row = r + 1 + i
    style_data_cell(ws1, row, 1).value = outcome
    style_data_cell(ws1, row, 2, "center").value = plans
    c = style_data_cell(ws1, row, 3, "right")
    c.value = parts
    c.number_format = num_fmt

# Total row
row += 1
total_cell = style_data_cell(ws1, row, 1)
total_cell.value = "Total"
total_cell.font = Font(bold=True)
c = style_data_cell(ws1, row, 2, "center")
c.value = 137
c.font = Font(bold=True)
c = style_data_cell(ws1, row, 3, "right")
c.value = 1695359
c.number_format = num_fmt
c.font = Font(bold=True)

auto_width(ws1)


# ===== SHEET 2: Likely Late Filers =====
ws2 = wb.create_sheet("Likely Late Filers")
ws2.sheet_properties.tabColor = "FFC000"

ws2.cell(row=1, column=1,
         value="Likely Late Filers (4 Plans)").font = title_font
ws2.merge_cells("A1:E1")
ws2.cell(row=2, column=1,
         value="Still Code 3 in 2023 filing; no 2024 filing yet. Almost certainly still active.").font = Font(
    italic=True, size=10)

headers = ["Sponsor", "Tracking ID", "EIN", "Plan Number", "Participants (2023)", "Notes"]
r = 4
for i, h in enumerate(headers, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, 6)

late_data = [
    ("General Motors LLC", "270383222-003", "270383222", "003", 335791,
     "Largest DB plan in the U.S.; substitute mortality every year since at least 2017"),
    ("General Motors LLC", "270383222-016", "270383222", "016", 33036,
     "Second GM plan; same filing pattern as 003"),
    ("Michelin North America, Inc.", "111724631-010", "111724631", "010", 19397,
     "Large manufacturing plan"),
    ("Komatsu Mining Corp.", "391566457-008", "391566457", "008", 11088,
     "Mining/heavy equipment manufacturer"),
]
for i, (sponsor, tid, ein, pn, parts, notes) in enumerate(late_data):
    row = 5 + i
    style_data_cell(ws2, row, 1).value = sponsor
    style_data_cell(ws2, row, 2, "center").value = tid
    style_data_cell(ws2, row, 3, "center").value = ein
    style_data_cell(ws2, row, 4, "center").value = pn
    c = style_data_cell(ws2, row, 5, "right")
    c.value = parts
    c.number_format = num_fmt
    style_data_cell(ws2, row, 6).value = notes

r = 10
ws2.cell(row=r, column=1,
         value="If these 4 plans file 2024 with Code 3, adjusted 2024 count = 59 plans / ~1,349,000 participants").font = Font(
    italic=True, size=10)

auto_width(ws2)


# ===== SHEET 3: Switched to Prescribed =====
ws3 = wb.create_sheet("Switched to Prescribed")
ws3.sheet_properties.tabColor = "70AD47"

ws3.cell(row=1, column=1,
         value="Plans That Switched from Substitute to Prescribed Tables (38 Plans, Still Active)").font = title_font
ws3.merge_cells("A1:F1")

headers = ["Sponsor", "Tracking ID", "EIN", "Plan Number", "Year Left Substitute", "Participants (2024)"]
r = 3
for i, h in enumerate(headers, 1):
    ws3.cell(row=r, column=i, value=h)
style_header_row(ws3, r, 6)

switched = [
    ("Georgia-Pacific LLC", "930432081-070", "930432081", "070", 2023, 35298),
    ("Weyerhaeuser Company", "910470860-002", "910470860", "002", 2019, 20057),
    ("Georgia-Pacific LLC", "930432081-046", "930432081", "046", 2023, 13368),
    ("Koch Companies, LLC", "992447784-001", "992447784", "001", 2023, 12355),
    ("OfficeMax Inc. (subsidiary)", "820100960-005", "820100960", "005", 2023, 12102),
    ("Ball Corporation", "350160610-001", "350160610", "001", 2020, 10345),
    ("Howmet Aerospace Inc.", "250317820-001", "250317820", "001", 2021, 9786),
    ("Arconic Corporation", "842745636-004", "842745636", "004", 2021, 8609),
    ("O-I Glass, Inc.", "222781933-002", "222781933", "002", 2022, 6344),
    ("ASARCO LLC", "810666284-008", "810666284", "008", 2023, 6259),
    ("Resolute FP US Inc.", "620721803-004", "620721803", "004", 2023, 4647),
    ("O-I Glass, Inc.", "222781933-001", "222781933", "001", 2022, 4548),
    ("Ball Corporation", "350160610-035", "350160610", "035", 2023, 4387),
    ("Alcoa USA Corp.", "371808900-002", "371808900", "002", 2022, 4314),
    ("Arconic Corporation", "842745636-003", "842745636", "003", 2021, 4303),
    ("Ardagh Glass Inc.", "351958205-008", "351958205", "008", 2023, 4051),
    ("Legacy Vulcan LLC", "208579133-010", "208579133", "010", 2023, 3366),
    ("Hexion Inc.", "130511250-002", "130511250", "002", 2022, 3049),
    ("INV Management Services, LLC", "851507460-001", "851507460", "001", 2023, 2772),
    ("ASARCO LLC", "810666284-007", "810666284", "007", 2023, 2751),
    ("Alcoa USA Corp.", "371808900-001", "371808900", "001", 2022, 2630),
    ("Suriname Aluminum Company LLC", "980150255-037", "980150255", "037", 2022, 2514),
    ("Legacy Vulcan LLC", "208579133-020", "208579133", "020", 2023, 2493),
    ("Molex LLC", "362369491-002", "362369491", "002", 2023, 2300),
    ("ATI Inc.", "251792394-001", "251792394", "001", 2021, 1996),
    ("Comau LLC", "382296242-001", "382296242", "001", 2019, 1973),
    ("Alcoa USA Corp.", "371808900-037", "371808900", "037", 2022, 1938),
    ("Suriname Aluminum Company, LLC", "980150255-003", "980150255", "003", 2022, 1099),
    ("Flint Hills Resources Pine Bend, LLC", "611603905-001", "611603905", "001", 2023, 997),
    ("Peabody Holding Company, LLC", "742666822-004", "742666822", "004", 2022, 968),
    ("Sonoco Products Company", "570248420-001", "570248420", "001", 2020, 889),
    ("Flowers Foods, Inc.", "582582379-003", "582582379", "003", 2019, 769),
    ("Buzzi Unicem USA Inc.", "233022369-010", "233022369", "010", 2020, 687),
    ("Trivium Packaging USA Inc.", "251864585-011", "251864585", "011", 2019, 345),
    ("Howmet Corporation", "132838093-009", "132838093", "009", 2021, 343),
    ("RMI Titanium Co., LLC", "310875005-001", "310875005", "001", 2021, 300),
    ("Motiva Enterprises LLC", "760262490-006", "760262490", "006", 2021, 165),
    ("FHR Peru Holding Company, LLC", "842606672-001", "842606672", "001", 2023, 72),
]
for i, (sponsor, tid, ein, pn, left_yr, parts) in enumerate(switched):
    row = 4 + i
    style_data_cell(ws3, row, 1).value = sponsor
    style_data_cell(ws3, row, 2, "center").value = tid
    style_data_cell(ws3, row, 3, "center").value = ein
    style_data_cell(ws3, row, 4, "center").value = pn
    style_data_cell(ws3, row, 5, "center").value = left_yr
    c = style_data_cell(ws3, row, 6, "right")
    c.value = parts
    c.number_format = num_fmt

auto_width(ws3)


# ===== SHEET 4: Terminated, PRT & Merged =====
ws4 = wb.create_sheet("Terminated, PRT & Merged")
ws4.sheet_properties.tabColor = "FF0000"

ws4.cell(row=1, column=1,
         value="Plans No Longer Filing (Terminated, PRT, Merged)").font = title_font
ws4.merge_cells("A1:H1")


def write_section(ws, start_row, title, headers, data, has_dollar_col=False):
    """Write a titled section with header row and data rows."""
    ws.cell(row=start_row, column=1, value=title).font = section_font
    hr = start_row + 1
    ncols = len(headers)
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header_row(ws, hr, ncols)

    for i, row_data in enumerate(data):
        row = hr + 1 + i
        for j, val in enumerate(row_data):
            col = j + 1
            if col == 1:
                align = "left"
            elif col == ncols and has_dollar_col:
                align = "right"
            elif col == ncols - (1 if has_dollar_col else 0):
                align = "right"
            else:
                align = "center"
            c = style_data_cell(ws, row, col, align)
            c.value = val
            if col == ncols - (1 if has_dollar_col else 0):
                c.number_format = num_fmt
            if has_dollar_col and col == ncols:
                c.number_format = dollar_fmt

    return hr + 1 + len(data) + 1  # next available row


# Terminated (no PRT)
terminated = [
    ("ThyssenKrupp North America, LLC", "222393554-009", "222393554", "009", 2019, 2019, 1936),
    ("Trane Technologies Company LLC", "135156640-050", "135156640", "050", 2020, 2020, 2593),
    ("Legacy Vulcan LLC", "208579133-030", "208579133", "030", 2020, 2020, 4494),
    ("Divested Litchfield Park Property, Inc.", "510304856-003", "510304856", "003", 2020, 2020, 25),
    ("ThyssenKrupp Materials NA, Inc.", "380445860-002", "380445860", "002", 2020, 2020, 75),
    ("Handy & Harman Ltd.", "133768097-001", "133768097", "001", 2021, 2021, 5671),
    ("ThyssenKrupp Crankshaft Company, LLC", "204023938-001", "204023938", "001", 2021, 2021, 139),
    ("ThyssenKrupp Industrial Solutions (USA), Inc.", "391858155-003", "391858155", "003", 2021, 2021, 219),
    ("ThyssenKrupp Rothe Erde USA Inc.", "340906272-005", "340906272", "005", 2021, 2022, 165),
    ("ThyssenKrupp Rothe Erde USA Inc.", "340906272-007", "340906272", "007", 2021, 2022, 266),
    ("Whirlpool Corporation", "381490038-125", "381490038", "125", 2022, 2022, 12124),
    ("TEK L.P. and Kote L.P.", "363525438-333", "363525438", "333", 2022, 2022, 552),
    ("FCA US LLC", "270187394-038", "270187394", "038", 2023, 2023, 6081),
    ("FCA US LLC", "270187394-043", "270187394", "043", 2023, 2023, 6719),
    ("The Goodyear Tire & Rubber Company", "340253240-017", "340253240", "017", 2023, 2023, 603),
    ("Cleveland-Cliffs Inc.", "341464672-003", "341464672", "003", 2023, 2023, 2759),
    ("Appalachia Holding Company", "540295165-001", "540295165", "001", 2023, 2023, 6813),
    ("Alpha Natural Resources, LLC", "562298262-003", "562298262", "003", 2023, 2023, 1521),
    ("Alpha Natural Resources, LLC", "562298262-004", "562298262", "004", 2023, 2023, 996),
]
r = 3
next_r = write_section(ws4, r,
                       "Terminated (Final Filing, No PRT) - 19 Plans",
                       ["Sponsor", "Tracking ID", "EIN", "Plan Number", "Last Sub Year", "Last Filing Year",
                        "Participants"],
                       terminated)

# Terminated with PRT
prt_term = [
    ("RMI Titanium Co., LLC", "310875005-004", "310875005", "004", 2021, 2021, 921, 71672000),
]
next_r = write_section(ws4, next_r,
                       "Terminated with PRT - 1 Plan",
                       ["Sponsor", "Tracking ID", "EIN", "Plan Number", "Last Sub Year", "Last Filing Year",
                        "Participants", "PRT Amount"],
                       prt_term, has_dollar_col=True)

# PRT Activity
prt_activity = [
    ("Howmet Corporation", "132838093-016", "132838093", "016", 2021, 2021, 4655, 15753670),
    ("Huck International, Inc.", "330483331-002", "330483331", "002", 2021, 2021, 612, 4269000),
    ("GKN America Corp.", "510282662-001", "510282662", "001", 2022, 2022, 4143, 52471670),
    ("FirstGroup Services, Inc.", "863006037-001", "863006037", "001", 2023, 2023, 5524, 69718176),
    ("Verizon Business Global LLC", "900357488-008", "900357488", "008", 2023, 2023, 824, 25861290),
]
next_r = write_section(ws4, next_r,
                       "PRT Activity (Winding Down, No Final Filing) - 5 Plans",
                       ["Sponsor", "Tracking ID", "EIN", "Plan Number", "Last Sub Year", "Last Filing Year",
                        "Participants", "PRT Amount"],
                       prt_activity, has_dollar_col=True)

# Plan Mergers
mergers = [
    ("Sonoco Products Company", "570248420-002", "570248420", "002", 2019, 2019, 8398, 1),
    ("Arconic Inc.", "250317820-002", "250317820", "002", 2021, 2021, 23282, 2),
    ("Huntington Ingalls Industries, Inc.", "900607005-269", "900607005", "269", 2021, 2021, 482, 4),
    ("ATI Inc.", "251792394-200", "251792394", "200", 2022, 2022, 2956, 1),
    ("ThyssenKrupp North America, LLC", "222393554-001", "222393554", "001", 2023, 2023, 3114, 1),
    ("Ball Corporation", "350160610-039", "350160610", "039", 2023, 2023, 5601, 2),
]
next_r = write_section(ws4, next_r,
                       "Plan Mergers (EIN Still Has Active Plans in 2024) - 6 Plans",
                       ["Sponsor", "Tracking ID", "EIN", "Plan Number", "Last Sub Year", "Last Filing Year",
                        "Participants", "Other EIN Plans in 2024"],
                       mergers)

# No Longer Filing
unclear = [
    ("Flint Hills Resources Peru, LLC", "471800638-001", "471800638", "001", 2019, 2019, 72),
    ("Motiva Chemicals LLC", "611603903-006", "611603903", "006", 2019, 2019, 202),
    ("INVISTA S.A.R.L.", "980196061-001", "980196061", "001", 2019, 2019, 2982),
    ("ThyssenKrupp Materials NA, Inc.", "380445860-001", "380445860", "001", 2020, 2020, 676),
    ("Greyhound Lines, Inc.", "860572343-003", "860572343", "003", 2020, 2020, 6673),
    ("The Western Union Company", "204531180-002", "204531180", "002", 2020, 2021, 5512),
    ("Peabody Investments Corp.", "200480084-005", "200480084", "005", 2022, 2022, 4151),
    ("Koch Industries, Inc.", "480484227-001", "480484227", "001", 2022, 2022, 14137),
    ("BlueLinx Corporation", "770627351-001", "770627351", "001", 2022, 2022, 2137),
]
write_section(ws4, next_r,
              "No Longer Filing (Status Unclear) - 9 Plans",
              ["Sponsor", "Tracking ID", "EIN", "Plan Number", "Last Sub Year", "Last Filing Year", "Participants"],
              unclear)

auto_width(ws4)


# ===== SHEET 5: Still Active (55 Plans) — Full 2009-2024 History =====
ws5 = wb.create_sheet("Still Active (55)")
ws5.sheet_properties.tabColor = "00B050"

ws5.cell(row=1, column=1,
         value="Plans Still Using Substitute Mortality in 2024 (55 Plans) - Full History 2009-2024").font = title_font
ws5.merge_cells("A1:V1")
ws5.cell(row=2, column=1,
         value="First Sub Year = first year plan filed as Code 3 (Substitute Mortality) in Schedule SB data (2009-2024). "
               "'2009 or earlier' means plan was already Code 3 when DOL EFAST2 data begins.").font = Font(italic=True, size=10)

YEARS = list(range(2009, 2025))
headers = ["Sponsor", "Tracking ID", "EIN", "Plan Number",
           "Participants (2024)", "First Sub Year"] + [str(y) for y in YEARS]
r = 4
for i, h in enumerate(headers, 1):
    ws5.cell(row=r, column=i, value=h)
style_header_row(ws5, r, len(headers))

# Code color fills
fill_sub = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
fill_prescribed = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
fill_none = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
fill_nan = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Full 2009-2024 data (sorted by participants descending)
# Format: (Sponsor, TID, EIN, PN, Participants, FirstSub, code_2009..code_2024)
active_plans = [
    ("Ford Motor Company", "380549190-001", "380549190", "001", 145606, 2015,
     "2","2","2","2","2","2","3","3","3","3","3","3","3","3","3","3"),
    ("Verizon Communications Inc.", "232259884-016", "232259884", "016", 112363, 2019,
     "-","-","-","-","-","-","-","2","2","2","3","3","3","3","3","3"),
    ("FCA US LLC", "270187394-005", "270187394", "005", 103598, 2012,
     "-","2","2","3","3","3","3","3","3","3","3","3","3","3","3","3"),
    ("Verizon Corporate Services Group Inc.", "131675522-001", "131675522", "001", 62044, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Ford Motor Company", "380549190-002", "380549190", "002", 57113, 2015,
     "2","2","2","2","2","2","3","3","3","3","3","3","3","3","3","3"),
    ("Eaton Corporation", "340196300-029", "340196300", "029", 46522, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Corning Incorporated", "160393470-001", "160393470", "001", 33672, 2021,
     "2","2","2","2","2","2","2","2","2","2","2","2","3","3","3","3"),
    ("Whirlpool Corporation", "381490038-107", "381490038", "107", 29187, 2019,
     "-","-","-","-","-","-","-","-","-","-","3","3","3","3","3","3"),
    ("United States Steel Corporation", "251897152-001", "251897152", "001", 24627, "2009 or earlier",
     "3","3","3","3","3","3","3","3","3","3","3","3","3","3","3","3"),
    ("The Goodyear Tire & Rubber Company", "340253240-001", "340253240", "001", 22650, "2009 or earlier",
     "3","3","3","3","3","3","3","3","3","3","3","3","3","3","3","3"),
    ("Trane Technologies Company LLC", "135156640-008", "135156640", "008", 19673, 2023,
     "-","-","-","-","-","-","-","-","-","-","-","-","-","-","3","3"),
    ("Olin Corporation", "131872319-002", "131872319", "002", 19133, 2019,
     "2","2","2","2","2","2","2","-","2","2","3","3","3","3","3","3"),
    ("Huntington Ingalls Industries, Inc.", "900607005-101", "900607005", "101", 17562, 2019,
     "-","-","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Crown Cork & Seal Company, Inc.", "231526444-001", "231526444", "001", 15568, 2020,
     "2","2","2","2","2","2","2","2","2","2","2","3","3","3","3","3"),
    ("Huntington Ingalls Industries, Inc.", "900607005-305", "900607005", "305", 14624, 2019,
     "-","-","2","2","2","-","2","2","2","2","3","3","3","3","3","3"),
    ("HM US Services, LLC", "814086708-003", "814086708", "003", 14076, 2021,
     "-","-","-","-","-","-","-","-","2","2","2","2","3","3","3","3"),
    ("Trane U.S. Inc.", "250900465-023", "250900465", "023", 13770, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("The Goodyear Tire & Rubber Company", "340253240-002", "340253240", "002", 13016, 2018,
     "2","2","2","2","2","-","2","2","2","3","3","3","3","3","3","3"),
    ("FCA US LLC", "270187394-004", "270187394", "004", 12560, 2012,
     "2","2","2","3","3","3","3","3","3","3","3","3","3","3","3","3"),
    ("Newell Operating Company", "361953130-001", "361953130", "001", 11941, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","2","2","2","3"),
    ("Huntington Ingalls Industries, Inc.", "900607005-100", "900607005", "100", 11906, 2019,
     "-","-","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Cleveland-Cliffs Steel Corporation", "311267098-003", "311267098", "003", 11463, 2016,
     "2","-","2","2","2","2","2","3","3","3","3","2","2","2","3","3"),
    ("Daimler Truck North America", "930790608-003", "930790608", "003", 11257, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Huntington Ingalls Industries, Inc.", "900607005-041", "900607005", "041", 11188, 2019,
     "-","-","2","2","2","-","2","2","2","2","3","3","3","3","3","3"),
    ("Cleveland-Cliffs Steel LLC", "710871875-009", "710871875", "009", 11173, "2009 or earlier",
     "3","3","3","3","3","3","3","3","3","3","3","3","3","3","3","3"),
    ("Quad/Graphics Printing LLC", "522009152-001", "522009152", "001", 10875, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("The Brink's Company", "541317776-003", "541317776", "003", 10340, 2019,
     "-","-","-","-","-","-","-","-","2","2","3","3","nan","3","3","3"),
    ("Alpha Natural Resources, LLC", "562298262-005", "562298262", "005", 9353, 2024,
     "-","-","-","-","-","-","-","-","-","-","-","-","-","-","-","3"),
    ("FCA US LLC", "270187394-007", "270187394", "007", 7640, 2012,
     "-","2","2","3","3","3","3","3","3","3","3","3","3","3","3","3"),
    ("The Goodyear Tire & Rubber Company", "340253240-010", "340253240", "010", 5938, 2018,
     "2","2","2","2","2","-","2","2","2","3","3","3","3","3","3","3"),
    ("Handy & Harman Ltd.", "133768097-002", "133768097", "002", 5825, 2019,
     "-","-","-","-","-","-","-","2","2","2","3","3","3","3","3","3"),
    ("Dana Limited", "261318190-148", "261318190", "148", 4376, 2019,
     "-","-","-","-","-","-","-","-","2","2","3","3","3","3","3","3"),
    ("Mueller Group, LLC", "371387813-005", "371387813", "005", 4374, 2019,
     "-","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Cleveland-Cliffs Steel LLC", "710871875-010", "710871875", "010", 4182, 2018,
     "-","-","-","-","-","-","-","-","2","3","3","3","3","3","3","3"),
    ("Daimler Truck North America", "930790608-001", "930790608", "001", 3622, 2019,
     "nan","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("JPS Industries Holdings LLC", "474722905-001", "474722905", "001", 3547, 2019,
     "-","-","-","-","-","-","2","2","2","2","3","3","3","3","3","3"),
    ("Dana Limited", "261318190-003", "261318190", "003", 3438, 2019,
     "2","2","2","2","2","2","2","2","2","-","3","3","3","3","3","3"),
    ("HarbisonWalker International, Inc.", "431680037-430", "431680037", "430", 3250, 2021,
     "-","-","-","-","-","-","-","2","2","2","2","2","3","3","3","3"),
    ("Instant Brands Holdings, Inc.", "161403318-004", "161403318", "004", 2990, 2021,
     "2","2","2","2","2","2","2","2","2","2","2","2","3","3","3","3"),
    ("The Cleveland-Cliffs Iron Company", "340677332-004", "340677332", "004", 2891, 2023,
     "-","-","-","-","-","-","-","-","-","2","2","2","2","2","3","3"),
    ("HarbisonWalker International, Inc.", "431680037-120", "431680037", "120", 2332, 2021,
     "-","-","-","-","-","-","-","2","2","2","2","2","3","3","3","3"),
    ("Trane Technologies Company LLC", "135156640-001", "135156640", "001", 2187, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Detroit Diesel Corporation", "382772023-002", "382772023", "002", 2128, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Cooper Tire & Rubber Company LLC", "344297750-007", "344297750", "007", 1828, 2023,
     "2","2","2","2","2","2","2","2","2","2","2","2","2","2","3","3"),
    ("Cleveland-Cliffs Hibbing Management LLC", "834490897-001", "834490897", "001", 1660, 2023,
     "-","-","-","-","-","-","-","-","-","-","2","2","2","2","3","3"),
    ("Buzzi Unicem USA Inc.", "233022369-021", "233022369", "021", 1612, 2018,
     "2","2","2","2","2","2","2","2","2","3","3","3","3","3","3","3"),
    ("Pilkington North America, Inc.", "341506654-030", "341506654", "030", 1599, 2020,
     "2","2","2","2","2","2","2","2","2","2","2","3","3","3","3","3"),
    ("Fairfield Manufacturing Company Inc.", "630500160-002", "630500160", "002", 1488, 2022,
     "2","2","nan","2","2","2","2","2","2","2","2","2","2","3","3","3"),
    ("Pilkington North America, Inc.", "341506654-040", "341506654", "040", 1425, 2020,
     "2","2","2","2","2","2","2","2","2","2","2","3","3","3","3","3"),
    ("Cooper Tire & Rubber Company LLC", "344297750-001", "344297750", "001", 1290, 2023,
     "2","2","2","2","2","2","2","2","2","2","2","2","2","2","3","3"),
    ("Olin Corporation", "131872319-081", "131872319", "081", 1158, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("ThyssenKrupp North America, LLC", "222393554-003", "222393554", "003", 1039, 2021,
     "-","-","-","-","-","-","-","-","-","-","-","-","3","2","3","3"),
    ("Detroit Diesel Corporation", "382772023-001", "382772023", "001", 947, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Dana Limited", "261318190-015", "261318190", "015", 283, 2019,
     "2","2","2","2","2","2","2","2","2","2","3","3","3","3","3","3"),
    ("Crown Cork & Seal Company, Inc.", "231526444-150", "231526444", "150", 144, 2020,
     "-","-","2","2","2","2","2","2","2","2","2","3","3","3","3","3"),
]

for i, plan in enumerate(active_plans):
    row = r + 1 + i
    sponsor, tid, ein, pn, parts, first_sub = plan[:6]
    codes = plan[6:]  # 16 year codes (2009-2024)

    style_data_cell(ws5, row, 1).value = sponsor
    style_data_cell(ws5, row, 2, "center").value = tid
    style_data_cell(ws5, row, 3, "center").value = ein
    style_data_cell(ws5, row, 4, "center").value = pn
    c = style_data_cell(ws5, row, 5, "right")
    c.value = parts
    c.number_format = num_fmt
    c = style_data_cell(ws5, row, 6, "center")
    c.value = first_sub if isinstance(first_sub, str) else first_sub

    # Year columns with color coding
    for j, code in enumerate(codes):
        col = 7 + j  # columns 7-22 = 2009-2024
        c = style_data_cell(ws5, row, col, "center")
        if code == "3":
            c.value = "Sub (3)"
            c.fill = fill_sub
            c.font = Font(color="FFFFFF", bold=True)
        elif code == "2":
            c.value = "Rx (2)"
            c.fill = fill_prescribed
        elif code == "1":
            c.value = "Comb (1)"
            c.fill = fill_prescribed
        elif code == "-":
            c.value = ""
            c.fill = fill_none
        elif code == "nan":
            c.value = "?"
            c.fill = fill_nan
        else:
            c.value = code

auto_width(ws5, min_width=10)
# Set year columns to consistent width
for col_idx in range(7, 7 + len(YEARS)):
    ws5.column_dimensions[get_column_letter(col_idx)].width = 10
ws5.freeze_panes = "G5"

# Save
out_path = "substitute_mortality_analysis_v2.xlsx"
wb.save(out_path)
print(f"Saved {out_path}")
